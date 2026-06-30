import re
import pandas as pd
import logging
import hashlib
from datetime import datetime
from reviews.models import Review, ReviewAnalysis, ReviewSource, VeratourReport
from django.utils import timezone
from transformers import pipeline

logger = logging.getLogger(__name__)

# Global variable for the sentiment pipeline
_sentiment_pipeline = None

def get_sentiment_pipeline():
    global _sentiment_pipeline
    if _sentiment_pipeline is None:
        try:
            # 100% Local Sentiment Analysis for Italian
            _sentiment_pipeline = pipeline('sentiment-analysis', model='MilaNLProc/feel-it-italian-sentiment')
        except Exception as e:
            logger.error(f"Failed to load sentiment pipeline: {e}")
    return _sentiment_pipeline

def extract_date_from_text(text):
    """
    Extracts date from text using Regex.
    Supports YYYY-MM-DD (Veratour new format) and DD/MM/YYYY (Legacy).
    """
    if not isinstance(text, str):
        return None

    # New Veratour Pattern: CODICE - DATA_INIZIO -> DATA_FINE (es: 2024GR000598 - 2024-05-25 -> 2024-06-01)
    # We need the second date (DATA_FINE)
    veratour_pattern = r'\d{4}-\d{2}-\d{2}'
    dates = re.findall(veratour_pattern, text)
    if len(dates) >= 2:
        try:
            dt = datetime.strptime(dates[1], '%Y-%m-%d')
            return timezone.make_aware(dt)
        except ValueError:
            pass
    elif len(dates) == 1:
        try:
            dt = datetime.strptime(dates[0], '%Y-%m-%d')
            return timezone.make_aware(dt)
        except ValueError:
            pass

    # Legacy support: 01/01/2024 or 1/1/2024
    legacy_pattern = r'(\d{1,2}/\d{1,2}/\d{4})'
    match = re.search(legacy_pattern, text)
    if match:
        date_str = match.group(1)
        for fmt in ('%d/%m/%Y', '%d/%m/%y'):
            try:
                dt = datetime.strptime(date_str, fmt)
                return timezone.make_aware(dt)
            except ValueError:
                continue
    return None

def get_sentiment_score_and_label(text):
    """
    Analyzes text using a LOCAL Italian model and returns a score (1-10) and a label.
    Implementation of Multi-Topic weighting and Weighted Relevance.
    Mapping to 5 clusters: 1-4 (Negative), 5-6 (Neutral), 7-10 (Positive)
    """
    if not text or not text.strip():
        return 6, 'neutral', 0.0

    # Split text by common separators for Multi-Topic analysis
    # Using points, commas, semicolons, etc. as requested.
    segments = [s.strip() for s in re.split(r'[.;!?,]', text) if s.strip()]
    if not segments:
        segments = [text.strip()]

    pipe = get_sentiment_pipeline()
    if not pipe:
        return 6, 'neutral', 0.0

    segment_results = []

    for segment in segments:
        try:
            # Analyze each segment
            res = pipe(segment[:512])[0]
            label = res['label']
            conf = res['score']

            # Map to 1-10
            if conf < 0.6:
                seg_score = 6 # Neutral
            elif label == 'positive':
                if conf > 0.9: seg_score = 10
                elif conf > 0.75: seg_score = 8
                else: seg_score = 7
            else: # negative
                if conf > 0.9: seg_score = 2
                elif conf > 0.75: seg_score = 4
                else: seg_score = 5 # Weak negative -> Neutralish

            # Identify department/topic for weighting
            weight = 1.0
            lower_seg = segment.lower()
            for topic, topic_weight in TOPIC_RELEVANCE.items():
                if topic in lower_seg:
                    weight = max(weight, topic_weight)

            # If no topic found, check departments
            if weight == 1.0:
                for dept, keywords in DEPARTMENT_KEYWORDS.items():
                    if any(k in lower_seg for k in keywords):
                        weight = max(weight, RELEVANCE_WEIGHTS.get(dept, 1.0))

            segment_results.append({
                'score': seg_score,
                'weight': weight,
                'polarity': conf if label == 'positive' else -conf
            })
        except Exception as e:
            logger.error(f"Segment sentiment analysis error: {e}")
            continue

    if not segment_results:
        return 6, 'neutral', 0.0

    # Calculate Weighted Average
    total_weighted_score = sum(r['score'] * r['weight'] for r in segment_results)
    total_weighted_polarity = sum(r['polarity'] * r['weight'] for r in segment_results)
    total_weight = sum(r['weight'] for r in segment_results)

    final_score = round(total_weighted_score / total_weight)
    final_polarity = total_weighted_polarity / total_weight

    # Mapping: Negative: 1-4, Neutro: 5-6, Positivo: 7-10
    if final_score >= 7:
        final_label = 'positive'
    elif final_score <= 4:
        final_label = 'negative'
    else:
        final_label = 'neutral'

    return final_score, final_label, final_polarity

def parse_veratour_report(file_path, resort):
    """
    Parses the REPORT file to extract Total Guests, Date Range and Granular Stats.
    Uses a recursive pattern: Header -> Pos -> Neg -> Sub-items (--->).
    """
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    total_guests = 0
    start_date = None
    end_date = None
    report_data = {}

    # Extract Total Guests and Dates (Global Search)
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str):
                # Total Guests (B1 or similar)
                if "schede elaborate" in val.lower() or "totale persone" in val.lower() or "ospiti" in val.lower():
                    # Try next cell in same row
                    target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if target_cell.value:
                        try:
                            total_guests = int(re.search(r'(\d+)', str(target_cell.value)).group(1))
                        except (ValueError, AttributeError): pass

                # Dates
                dates = re.findall(r'(\d{2}/\d{2}/\d{4})', val)
                if len(dates) >= 2:
                    try:
                        start_date = datetime.strptime(dates[0], '%d/%m/%Y').date()
                        end_date = datetime.strptime(dates[1], '%d/%m/%Y').date()
                    except ValueError: pass

    # Granular Extraction using Department Headers (Yellow rows)
    # We identify departments by text and potentially row color if available,
    # but text patterns are safer if color info is lost.

    current_dept = None

    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        cell_b = ws.cell(row=row_idx, column=2)

        val_a = str(cell_a.value or "").strip()
        val_b = cell_b.value

        # Check if it's a Department Header (usually uppercase, or known list)
        depts = [
            "GENERAL", "CAMERA", "RISTORAZIONE", "SERVIZI HOTEL",
            "PULIZIA AREE COMUNI", "ASSISTENZA VERATOUR", "ANIMAZIONE",
            "SPORT", "MINI CLUB", "ESCURSIONI", "TRASPORTO AEREO", "TRASFERIMENTI"
        ]

        found_dept = None
        for d in depts:
            if d in val_a.upper():
                found_dept = d
                break

        if found_dept:
            current_dept = found_dept
            report_data[current_dept] = {
                "positive": 0,
                "negative": 0,
                "sub_items": {}
            }

            # Next row is Positive
            pos_row = ws.cell(row=row_idx + 1, column=2)
            if pos_row.value:
                try:
                    val = str(pos_row.value).replace('%', '').strip()
                    report_data[current_dept]["positive"] = float(val)
                except ValueError: pass

            # Next next row is Negative
            neg_row = ws.cell(row=row_idx + 2, column=2)
            if neg_row.value:
                try:
                    val = str(neg_row.value).replace('%', '').strip()
                    report_data[current_dept]["negative"] = float(val)
                except ValueError: pass
            continue

        # Check for sub-items (starting with --->)
        if current_dept and "--->" in val_a:
            label = val_a.replace("--->", "").strip()
            try:
                report_data[current_dept]["sub_items"][label] = int(val_b) if val_b else 0
            except (ValueError, TypeError):
                report_data[current_dept]["sub_items"][label] = 0

    return total_guests, start_date, end_date, report_data

DEPARTMENT_KEYWORDS = {
    "CAMERA": ["camera", "stanza", "comfort", "manutenzione", "letto", "bagno", "doccia", "materasso", "frigo bar", "aria condizionata", "balcone"],
    "RISTORAZIONE": ["ristorante", "cibo", "mangiare", "buffet", "qualità", "varietà", "ristorazione", "pasta", "carne", "pesce", "cena", "pranzo", "colazione", "maitre", "chef", "tavolo", "bevande", "vino"],
    "ANIMAZIONE": ["animazione", "balli", "spettacolo", "ragazzi", "team", "divertimento", "animatori", "intrattenimento", "capo animatore", "miniclub", "anfiteatro", "spettacoli"],
    "SERVIZI HOTEL": ["reception", "bar", "wi-fi", "internet", "accoglienza", "personale", "gentilezza"],
    "PULIZIA": ["pulizia", "pulito", "ordine", "igiene", "sporco", "aree comuni", "piscina", "spiaggia", "giardini"],
    "ASSISTENZA VERATOUR": ["assistenza", "veratour", "assistenti", "problemi", "cortesia", "efficienza"],
    "SPORT": ["sport", "tennis", "calcio", "palestra", "attrezzatura", "tornei"],
    "MINI CLUB": ["mini club", "bambini", "baby", "animazione piccoli"],
    "ESCURSIONI": ["escursione", "tour", "viaggio", "guida", "visita"],
    "SICUREZZA": ["sicurezza", "guardia", "pericolo", "sicuro"],
    "TRASPORTO AEREO": ["aereo", "volo", "aeroporto", "ritardo"],
    "TRASFERIMENTI": ["trasferimento", "bus", "navetta", "viaggio"],
}

# Relevance weights for sentiment calculation
RELEVANCE_WEIGHTS = {
    "PULIZIA": 1.5,
    "SICUREZZA": 1.5,
    "CAMERA": 1.0, # Comfort Camere / Manutenzione
    "RISTORAZIONE": 1.5, # Qualità del cibo (if it's Ristorazione it's high)
    "ANIMAZIONE": 1.0,
    "MINI CLUB": 1.0,
    "SPORT": 1.0,
    "SERVIZI HOTEL": 0.5, # Wi-Fi, Bar are low, but Personnel is medium... let's refine
    "ASSISTENZA VERATOUR": 1.0,
    "ESCURSIONI": 0.5,
    "TRASPORTO AEREO": 0.5,
    "TRASFERIMENTI": 0.5,
}

# Explicit mapping for specific topics within departments if needed
TOPIC_RELEVANCE = {
    "pulizia": 1.5,
    "sicurezza": 1.5,
    "manutenzione": 1.5,
    "cibo": 1.5,
    "qualità": 1.5,
    "animazione": 1.0,
    "personale": 1.0,
    "assistenza": 1.0,
    "comfort": 1.0,
    "wi-fi": 0.5,
    "bar": 0.5,
    "escursioni": 0.5,
    "buffet": 0.5,
}

def process_vota_commenti(file_path, resort, progress_callback=None):
    """
    Parses VOTA_COMMENTI file and saves reviews.
    Handles the special Veratour block format.
    """
    # Read without header because it might be a mess
    df = pd.read_excel(file_path, header=None)
    source, _ = ReviewSource.objects.get_or_create(name='Veratour')

    # Flatten all data to find review blocks
    # According to requirements, each review block starts with a date pattern
    all_cells = df.values.flatten()
    total_cells = len(all_cells)

    processed_count = 0
    veratour_date_pattern = r'(\d{4}[A-Z]{2}\d+.*->.*\d{4}-\d{2}-\d{2})'

    for index, cell_value in enumerate(all_cells):
        if not isinstance(cell_value, str) or not cell_value.strip():
            continue

        # Skip headers like "VERACLUB AMASEA"
        if "VERACLUB" in cell_value.upper():
            continue

        # Try to extract date
        review_date = extract_date_from_text(cell_value)

        if review_date:
            # It's a review block!
            # The comment is everything in the cell AFTER the date pattern or on next lines
            # Clean up the text: remove the first line if it contains the pattern
            lines = cell_value.splitlines()
            comment_text = ""

            # If it's a multi-line cell, the first line is usually the metadata
            if len(lines) > 1:
                # Check if first line matches the metadata pattern
                if re.search(r'\d{4}-\d{2}-\d{2}', lines[0]):
                    comment_text = "\n".join(lines[1:]).strip()
                else:
                    comment_text = cell_value.strip()
            else:
                # Single line cell, try to remove the pattern from the beginning
                comment_text = re.sub(r'.*->.*\d{4}-\d{2}-\d{2}', '', cell_value).strip()

            if not comment_text:
                # If the cell only had the metadata, maybe the comment is in the next cell?
                # But user says 1:1 relation and "tutto il resto del testo in quel blocco è il commento"
                # So we proceed even if it's short, or skip if empty
                continue

            # Sentiment Analysis
            score, label, polarity = get_sentiment_score_and_label(comment_text)

            # Anomaly Detection: Report Rating >= 9 and IA Sentiment score <= 3
            # We don't have the "Report Rating" here yet as it's just the comment file.
            # But wait, in Veratour, the rating of the Review object IS the IA score?
            # No, looking at existing code:
            # review, created = Review.objects.update_or_create(..., defaults={'rating': score, ...})
            # This seems wrong for Veratour. Veratour reviews usually have a numeric rating in the report.
            # However, VOTA_COMMENTI only has comments.
            # Re-reading requirement: "Se un ospite assegna un voto numerico altissimo nel report ma scrive un commento ferocemente negativo..."
            # This implies we need to match the comment with the report entry.
            # But Veratour reports (quantitative) and comments are in separate files and don't seem to have a shared ID.
            # If they can't be matched 1:1, "Anomalia" might only apply to platforms where we HAVE both (Google, Booking).
            # OR we assume 'rating' in Review for Veratour comes from a different source if matched.
            # Given the current structure, let's stick to the 1-10 mapping.

            # Department Mapping (Cross-Analysis)
            matched_depts = []
            lower_text = comment_text.lower()
            for dept, keywords in DEPARTMENT_KEYWORDS.items():
                if any(k in lower_text for k in keywords):
                    matched_depts.append(dept)

            # Create Review
            # Use hash of text for better uniqueness if index is not reliable across cells
            text_hash = hashlib.md5(comment_text.encode()).hexdigest()[:10]
            review_id = f"VERA_{resort.id}_{review_date.strftime('%Y%m%d')}_{text_hash}"

            review, created = Review.objects.update_or_create(
                source=source,
                review_id=review_id,
                defaults={
                    'resort': resort,
                    'author': "Ospite Veratour",
                    'rating': score,
                    'text': comment_text,
                    'review_date': review_date
                }
            )

            ReviewAnalysis.objects.update_or_create(
                review=review,
                defaults={
                    'sentiment_score': polarity,
                    'sentiment_label': label,
                    'keywords': matched_depts, # Use keywords field to store matched departments for now
                    'is_anomaly': False, # For Veratour, we don't have the numeric vote in VOTA_COMMENTI
                }
            )

            processed_count += 1
            if progress_callback:
                progress_callback(processed_count, total_cells)

    return processed_count
